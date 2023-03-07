import datetime
import re
import icalendar
import openpyxl

#定义课程开始时间
StartTime=[datetime.time(8, 0, 0), datetime.time(8, 50, 0), datetime.time(9, 50, 0), 
           datetime.time(10, 40, 0), datetime.time(11, 30, 0), datetime.time(13, 00, 0), 
           datetime.time(13, 50, 0), datetime.time(14, 45, 0), datetime.time(15, 40, 0), 
           datetime.time(16, 35, 0), datetime.time(17, 25, 0), datetime.time(18, 30, 0), 
           datetime.time(19, 20, 0), datetime.time(20, 10, 0)]
#定义课程结束时间
EndTime=[datetime.time(8, 45, 0), datetime.time(9, 35, 0), datetime.time(10, 35, 0), 
         datetime.time(11, 25, 0), datetime.time(12, 15, 0), datetime.time(13, 45, 0), 
         datetime.time(14, 35, 0), datetime.time(15, 30, 0), datetime.time(16, 25, 0), 
         datetime.time(17, 20, 0), datetime.time(18, 10, 0), datetime.time(19, 15, 0), 
         datetime.time(20, 5, 0), datetime.time(20, 55, 0)]
#找到字符串中某个关键值的所有索引，存放在list(int)中
def GetElementIndex(char, string):
    return [idx.start() for idx in re.finditer(char, string)]
#将上课周数转为list(int)
def ChangeIntoList_int(s):
    ranges = re.findall(r'(\d+)-(\d+)', s)
    for start, end in ranges:
        s = s.replace(f'{start}-{end}', ','.join(map(str, range(int(start), int(end)+1))))
    return list(map(int, s.split(',')))

#获取学号，打开xlsx文件
userid=input('请输入学号，确保和xlsx文件名中的学号一致：')
WorkBook=openpyxl.load_workbook(filename=f"./学生个人课表_{userid}.xlsx")
Sheet=WorkBook.active

print("-------------------------")
print("您的信息为：")
print("课程表所属人：", end="")
print(Sheet['A1'].value.replace("北京邮电大学 ","").replace(" 学生个人课表",""))
StudentName=Sheet['A1'].value.replace("北京邮电大学 ","").replace(" 学生个人课表","")
print("学年：",end="")
print(Sheet['A2'].value[5:16])
SchoolYear=Sheet['A2'].value[5:16]

#输入学期的第一周的周一日期
#StartDate=datetime.date(2023, 2, 20)
StartDate=datetime.datetime.strptime(input("输入学期的第一周的周一的日期，以YYYY-MM-DD格式\n"), '%Y-%m-%d').date()
while StartDate.isoweekday() != 1:
    StartDate=datetime.datetime.strptime(input("日期并非周一！请以YYYY-MM-DD格式输入\n"), '%Y-%m-%d').date()
print("正在处理")


#制作部分
MyCalendar=icalendar.Calendar()
MyCalendar.add('PRODID', '-//MY_CALENDAR_PRODUCT//GL//')
MyCalendar.add('VERSION', '2.0') #固定属性，版本2.0
MyCalendar.add('CALSCALE', 'GREGORIAN') #公历
MyCalendar.add('METHOD', 'PUBLISH')
MyCalendar.add('X-WR-CALNAME', f'{SchoolYear}') #通用属性，日历名称，默认为学年
MyCalendar.add('X-WR-TIMEZONE', 'Asia/Shanghai') #通用属性，指定时区
MyCalendar.add('X-APPLE-CALENDAR-COLOR', '#E1FFFF') #Apple日历颜色，可自己更改，填入十六进制代码
for Column in range(2, 9):
    for Row in range(4, 18):
        CellBR=GetElementIndex("\n", Sheet.cell(row=Row, column=Column).value)
        for i in range(int(len(CellBR)/5)): #拆分课程、教师名字、上课周数、上课教室、上课节次
            Course=Sheet.cell(row=Row, column=Column).value[CellBR[5*i]+1:CellBR[5*i+1]]
            TeacherName=Sheet.cell(row=Row, column=Column).value[CellBR[5*i+1]+1:CellBR[5*i+2]]
            ClassWeeks=Sheet.cell(row=Row, column=Column).value[CellBR[5*i+2]+1:CellBR[5*i+3]]
            Classroom=Sheet.cell(row=Row, column=Column).value[CellBR[5*i+3]+1:CellBR[5*i+4]]
            if i==int(len(CellBR)/5)-1:
                LessonNum=Sheet.cell(row=Row, column=Column).value[CellBR[5*i+4]+1:]
            else:
                LessonNum=Sheet.cell(row=Row, column=Column).value[CellBR[5*i+4]+1:CellBR[5*i+5]]
            ListLessonNum=LessonNum.replace("[","").replace("]","").replace("节","").split("-")
            ListLessonNum=list(map(int, ListLessonNum))
            if (Row-3==ListLessonNum[0]): #是第一节课才添加，下一节跳过
                ListClassWeeks=ChangeIntoList_int(ClassWeeks.replace("[周]",""))
                for j in range(len(ListClassWeeks)):
                    MyEvent=icalendar.Event()
                    MyEvent.add('SUMMARY', Course+' '+Classroom) #事件名称：课程名加教室
                    MyEvent.add('DTSTAMP', datetime.datetime.today())
                    MyEvent.add('DTSTART', datetime.datetime.combine(StartDate+datetime.timedelta(weeks=ListClassWeeks[j]-1,days=Column-2), StartTime[ListLessonNum[0]-1]))
                    MyEvent.add('DTEND', datetime.datetime.combine(StartDate+datetime.timedelta(weeks=ListClassWeeks[j]-1,days=Column-2), EndTime[ListLessonNum[-1]-1]))
                    MyEvent.add('DESCRIPTION', TeacherName) #教师姓名写在备注里
                    MyAlarm=icalendar.Alarm() #添加提醒作为事件的附加属性
                    MyAlarm.add('TRIGGER', datetime.timedelta(minutes=-10)) #提前10分钟提醒
                    MyAlarm.add('ACTION', "DISPLAY") #通知提醒
                    MyAlarm.add('DESCRIPTION', Course) #提醒内容：课程名称
                    MyEvent.add_component(MyAlarm)
                    MyCalendar.add_component(MyEvent)
                    del MyAlarm, MyEvent
                del TeacherName, ClassWeeks, Classroom, LessonNum, Course, ListClassWeeks
        del CellBR
try:
    with open('TimeTable.ics', 'wb') as file:
        file.write(MyCalendar.to_ical())
        print('[Success]')
        del MyCalendar
except Exception:
    print("生成文件失败，请重试")